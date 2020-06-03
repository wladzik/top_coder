import re
import numpy
from concurrent.futures import ProcessPoolExecutor
from openpyxl import load_workbook
from time import perf_counter
from datetime import datetime

test_file = "E:/Загрузки/file.xlsx"
required_sheets = ["Fund_All_D", "FX_All_D"]

forex_prices_pattern = "INSERT INTO forexprices(`currencyCode`, `bbgTicker`, `date`, `price`) VALUES {arg};"
fund_data_pattern = "INSERT INTO funddata(`bbgTicker`, `isin`, `name`, `managementCompany`, `country`, `currency`, " \
                    "`fundObjective`, `totalAssets`, `minInvestmentUsd`, `lastPriceLc`, `date`) VALUES {arg};"
stock_pattern = "INSERT INTO stockdata(`stockCode`, `bbgTicker`, `company`, `gicsSector`, `gicsSectorName`, " \
                "`roundLotSize`, `marketCapitalization`, `lastPrice`, `adjustedClosingPrice`, " \
                "`avgDailyValueTraded5Day`, `avgDailyValueTraded20Day`, `trailingPE`, `priceBook`, " \
                "`trailing12MonthNetSales`, `pxToSalesRatio`, `trailing12MonthFreeCashFlow`, " \
                "`freeCashFlowYield`, `dividendYield`, `priceGrossCashFlow`, `enterpriseValueEbitda`, " \
                "`enterpriseValue`, `interestCoverageRatio`, `netDebtEbitda`, `currentSharesOutstanding`, " \
                "`totalValueOfSharesRepurchased`, `returnOnCommonEquity`, `grossProfit`, `totalAssets`, " \
                "`forwardPE`, `salesEstimatesFwd12MCurrent`, `salesEstimatesFwd12M1MAgo`, " \
                "`salesEstimatesFwd12M3MAgo`, `salesEstimatesFwd12M12MAgo`, `epsEstimatesFwd12MCurrent`, " \
                "`epsEstimatesFwd12M1MAgo`, `epsEstimatesFwd12M3MAgo`, `epsEstimatesFwd12M12MAgo`, " \
                "`opEstimatesFwd12MCurrent`, `opEstimatesFwd12M1MAgo`, `opEstimatesFwd12M3MAgo`, " \
                "`opEstimatesFwd12M12MAgo`, `bestAnalystRating`, `1MonthTotalReturnCurrent`, " \
                "`3MonthTotalReturnCurrent`, `6MonthTotalReturnCurrent`, `12MonthTotalReturnCurrent`, " \
                "`30DayIvolAt90Moneyness`, `30DayIvolAt95Moneyness`, `30DayIvolAt105Moneyness`, " \
                "`30DayIvolAt110Moneyness`, `30DayIvolAt100Moneyness`, `60DayIvolAt100Moneyness`, " \
                "`3MonthIvolAt100Moneyness`, `rsi14D`, `bollinger20DB`, `tickerAndExchangeCode`, `chineseName`, " \
                "`currencyFundamentalDataReported`, `date`) VALUES {arg};"


def none_filter(row):
    return [i for i in row if i is not None]   # do it only for header. may be improves performance


def parallel_worksheet(sheet_name):
    begin = perf_counter()
    wb = load_workbook(test_file, read_only=True, data_only=True, keep_links=False)
    ws = wb[sheet_name]
    data_from_sheet = [row for row in ws.iter_rows(values_only=True) if row[0] is not None]
    end = perf_counter()
    print("    {0} {1:.2f}s".format(sheet_name, end - begin))
    prepared_list = []
    if sheet_name == required_sheets[1]:
        transposed = numpy.transpose(data_from_sheet)
        len_tr = len(transposed[0])
        for x in transposed[1:]:
            n = 1
            while n < len_tr and x[0] is not None:
                prepared_list.append((x[0], x[0],
                                      datetime.strptime(transposed[0][n], "%d/%m/%Y").strftime("%Y-%m-%d"),
                                      float(x[n])))
                n += 1
    elif sheet_name == required_sheets[0]:
        for f in data_from_sheet[1:]:
            prepared_list.append((f[0], f[1] if len(f[1]) <= 10 else f[1][2:], *f[2:7], f[8],
                                  f[9] if type(f[9]) is int else None, f[10],
                                  datetime.strptime(f[12], "%d/%m/%Y").strftime("%Y-%m-%d")))
    else:
        for r in numpy.transpose(data_from_sheet)[1:]:
            prepared_list.append((r[1], r[0], *r[2:15], round(r[15], 2) if type(r[15]) is not str else None, *r[16:24],
                                  round(r[24], 2) if type(r[24]) is not str else None, r[25],
                                  round(r[26], 2) if type(r[26]) is not str else None,
                                  r[27], r[28], round(r[29], 2) if type(r[29]) is not str else None,
                                  *r[30:37], round(r[37], 2) if type(r[37]) is not str else None,
                                  *r[38:42], *r[45:57],
                                  round(r[57], 2), *r[58:61],
                                  datetime.strptime(r[64], "%d/%m/%Y").strftime("%Y-%m-%d")))
    return prepared_list, sheet_name


def parallel_read():
    print("Parallised Read")
    begin = perf_counter()
    wb = load_workbook(test_file, read_only=True, keep_links=False, data_only=True)
    print("    Workbook loaded {0:.2f}s".format(perf_counter() - begin))
    sheets = list(filter(lambda x: x in required_sheets or re.match(r"Stock_[A-Z][A-Z]_D", x), wb.sheetnames))
    string = []
    with ProcessPoolExecutor() as pool:
        for ws in pool.map(parallel_worksheet, sheets):
            if ws[1] == required_sheets[1]:
                string.append(forex_prices_pattern.format(arg=str(ws[0])[1:-1]).replace("None", "NULL"))
            elif ws[1] == required_sheets[0]:
                string.append(fund_data_pattern.format(arg=str(ws[0])[1:-1]).replace("None", "NULL"))
            else:
                string.append(stock_pattern.format(arg=str(ws[0])[1:-1]).replace("None", "NULL"))
    with open("insert.sql", 'w+', encoding="utf-8") as sql_file:
        sql_file.write("\n".join(string))
    end = perf_counter()
    print("    Total time {0:.2f}s".format(end - begin))


if __name__ == "__main__":
    parallel_read()
